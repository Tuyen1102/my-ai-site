import os
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pyodbc
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("mapping_output")
OUTPUT_DIR.mkdir(exist_ok=True)

TARGET_KEYWORDS = [
    "Cám 5a.10",
    "Cám 5a.14",
    "Cám 6a.10",
    "Cám 6a.14",
]

DB_NAMES = [
    "TTCO_THONGKE",
    "TTCO_QTHT",
    "TTCO_DHSX",
]


def norm(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return value


def qident(name: str) -> str:
    return "[" + name.replace("]", "]]") + "]"


def load_config() -> Dict[str, str]:
    load_dotenv()
    return {
        "DB_SERVER": os.getenv("DB_SERVER", "tuyenthancuaong.com.vn,52376").strip(),
        "DB_DRIVER": os.getenv("DB_DRIVER", "ODBC Driver 18 for SQL Server").strip(),
        "DB_AUTH": os.getenv("DB_AUTH", "windows").strip().lower(),
        "DB_USER": os.getenv("DB_USER", "").strip(),
        "DB_PASSWORD": os.getenv("DB_PASSWORD", "").strip(),
    }


def build_conn_str(config: Dict[str, str], db_name: str) -> str:
    base = (
        f"DRIVER={{{config['DB_DRIVER']}}};"
        f"SERVER={config['DB_SERVER']};"
        f"DATABASE={db_name};"
        "TrustServerCertificate=yes;"
        "Encrypt=no;"
    )

    if config["DB_AUTH"] == "sql":
        if not config["DB_USER"] or not config["DB_PASSWORD"]:
            raise RuntimeError("DB_AUTH=sql nhưng thiếu DB_USER hoặc DB_PASSWORD trong file .env")
        return base + f"UID={config['DB_USER']};PWD={config['DB_PASSWORD']};"

    return base + "Trusted_Connection=yes;"


def rows_to_dicts(cursor: pyodbc.Cursor) -> List[Dict[str, Any]]:
    cols = [c[0] for c in cursor.description]
    out = []
    for row in cursor.fetchall():
        out.append({cols[i]: norm(row[i]) for i in range(len(cols))})
    return out


def run_query(conn: pyodbc.Connection, sql: str, params: List[Any] | None = None) -> List[Dict[str, Any]]:
    cur = conn.cursor()
    cur.execute(sql, params or [])
    return rows_to_dicts(cur)


def get_latest_period(conn: pyodbc.Connection) -> Tuple[int, int]:
    sql = """
    SELECT TOP 1 NamHT, ThangHT
    FROM dbo.CDOTHAN
    WHERE NamHT IS NOT NULL AND ThangHT IS NOT NULL
    GROUP BY NamHT, ThangHT
    ORDER BY NamHT DESC, ThangHT DESC
    """
    data = run_query(conn, sql)
    if not data:
        raise RuntimeError("Không tìm thấy kỳ dữ liệu trong TTCO_THONGKE.dbo.CDOTHAN")
    return int(data[0]["NamHT"]), int(data[0]["ThangHT"])


def get_dmthan_columns(conn: pyodbc.Connection) -> List[Dict[str, Any]]:
    sql = """
    SELECT
        TABLE_CATALOG,
        TABLE_SCHEMA,
        TABLE_NAME,
        COLUMN_NAME,
        DATA_TYPE,
        ORDINAL_POSITION
    FROM INFORMATION_SCHEMA.COLUMNS
    WHERE TABLE_NAME = 'DMTHAN'
    ORDER BY TABLE_CATALOG, TABLE_SCHEMA, TABLE_NAME, ORDINAL_POSITION
    """
    return run_query(conn, sql)


def get_dmthan_all(conn: pyodbc.Connection) -> List[Dict[str, Any]]:
    sql = """
    SELECT *
    FROM dbo.DMTHAN
    ORDER BY MaThan
    """
    return run_query(conn, sql)


def get_dmthan_target_rows(conn: pyodbc.Connection) -> List[Dict[str, Any]]:
    where = " OR ".join(["TenThan LIKE ?" for _ in TARGET_KEYWORDS])
    params = [f"%{kw}%" for kw in TARGET_KEYWORDS]
    sql = f"""
    SELECT *
    FROM dbo.DMTHAN
    WHERE {where}
    ORDER BY TenThan, MaThan
    """
    return run_query(conn, sql, params)


def get_cdothan_target_rows(conn: pyodbc.Connection, nam: int, thang: int) -> List[Dict[str, Any]]:
    where = " OR ".join(["t.TenThan LIKE ?" for _ in TARGET_KEYWORDS])
    params = [nam, thang] + [f"%{kw}%" for kw in TARGET_KEYWORDS]

    sql = f"""
    SELECT
        c.NamHT,
        c.ThangHT,
        LTRIM(RTRIM(c.MaKho)) AS MaKho,
        COALESCE(k.TenKho, c.MaKho) AS TenKho,
        LTRIM(RTRIM(c.MaThan)) AS MaThan,
        t.TenThan,
        CAST(SUM(ISNULL(c.TonDK, 0)) AS float) AS TonDauKy,
        CAST(SUM(ISNULL(c.NhapTK, 0)) AS float) AS NhapTrongKy,
        CAST(SUM(ISNULL(c.XuatTK, 0)) AS float) AS XuatTrongKy,
        CAST(SUM(ISNULL(c.TonCK, 0)) AS float) AS TonCuoiKy
    FROM dbo.CDOTHAN c
    LEFT JOIN TTCO_QTHT.dbo.DMKHO k
        ON UPPER(LTRIM(RTRIM(c.MaKho))) = UPPER(LTRIM(RTRIM(k.MaKho)))
    LEFT JOIN TTCO_QTHT.dbo.DMTHAN t
        ON UPPER(LTRIM(RTRIM(c.MaThan))) = UPPER(LTRIM(RTRIM(t.MaThan)))
    WHERE c.NamHT = ?
      AND c.ThangHT = ?
      AND ({where})
    GROUP BY
        c.NamHT,
        c.ThangHT,
        LTRIM(RTRIM(c.MaKho)),
        COALESCE(k.TenKho, c.MaKho),
        LTRIM(RTRIM(c.MaThan)),
        t.TenThan
    ORDER BY
        COALESCE(k.TenKho, c.MaKho),
        t.TenThan,
        LTRIM(RTRIM(c.MaThan))
    """
    return run_query(conn, sql, params)


def get_cdothan_target_grouped_by_base(conn: pyodbc.Connection, nam: int, thang: int) -> List[Dict[str, Any]]:
    sql = """
    WITH src AS (
        SELECT
            c.NamHT,
            c.ThangHT,
            LTRIM(RTRIM(c.MaKho)) AS MaKho,
            COALESCE(k.TenKho, c.MaKho) AS TenKho,
            LTRIM(RTRIM(c.MaThan)) AS MaThan,
            t.TenThan,
            CASE
                WHEN t.TenThan LIKE N'Cám 5a.10%' THEN N'Cám 5a.10'
                WHEN t.TenThan LIKE N'Cám 5a.14%' THEN N'Cám 5a.14'
                WHEN t.TenThan LIKE N'Cám 6a.10%' THEN N'Cám 6a.10'
                WHEN t.TenThan LIKE N'Cám 6a.14%' THEN N'Cám 6a.14'
                ELSE t.TenThan
            END AS TenThanGop,
            c.TonDK,
            c.NhapTK,
            c.XuatTK,
            c.TonCK
        FROM dbo.CDOTHAN c
        LEFT JOIN TTCO_QTHT.dbo.DMKHO k
            ON UPPER(LTRIM(RTRIM(c.MaKho))) = UPPER(LTRIM(RTRIM(k.MaKho)))
        LEFT JOIN TTCO_QTHT.dbo.DMTHAN t
            ON UPPER(LTRIM(RTRIM(c.MaThan))) = UPPER(LTRIM(RTRIM(t.MaThan)))
        WHERE c.NamHT = ?
          AND c.ThangHT = ?
          AND (
              t.TenThan LIKE N'Cám 5a.10%'
              OR t.TenThan LIKE N'Cám 5a.14%'
              OR t.TenThan LIKE N'Cám 6a.10%'
              OR t.TenThan LIKE N'Cám 6a.14%'
          )
    )
    SELECT
        NamHT,
        ThangHT,
        MaKho,
        TenKho,
        TenThanGop,
        CAST(SUM(ISNULL(TonDK, 0)) AS float) AS TonDauKy,
        CAST(SUM(ISNULL(NhapTK, 0)) AS float) AS NhapTrongKy,
        CAST(SUM(ISNULL(XuatTK, 0)) AS float) AS XuatTrongKy,
        CAST(SUM(ISNULL(TonCK, 0)) AS float) AS TonCuoiKy,
        COUNT(*) AS SoDongGop,
        STRING_AGG(CAST(MaThan AS nvarchar(max)), N'; ') AS DanhSachMaThanGoc,
        STRING_AGG(CAST(TenThan AS nvarchar(max)), N'; ') AS DanhSachTenThanGoc
    FROM src
    GROUP BY
        NamHT,
        ThangHT,
        MaKho,
        TenKho,
        TenThanGop
    ORDER BY
        TenKho,
        TenThanGop
    """
    try:
        return run_query(conn, sql, [nam, thang])
    except Exception as exc:
        # Một số SQL Server cũ không hỗ trợ STRING_AGG.
        sql_fallback = """
        WITH src AS (
            SELECT
                c.NamHT,
                c.ThangHT,
                LTRIM(RTRIM(c.MaKho)) AS MaKho,
                COALESCE(k.TenKho, c.MaKho) AS TenKho,
                LTRIM(RTRIM(c.MaThan)) AS MaThan,
                t.TenThan,
                CASE
                    WHEN t.TenThan LIKE N'Cám 5a.10%' THEN N'Cám 5a.10'
                    WHEN t.TenThan LIKE N'Cám 5a.14%' THEN N'Cám 5a.14'
                    WHEN t.TenThan LIKE N'Cám 6a.10%' THEN N'Cám 6a.10'
                    WHEN t.TenThan LIKE N'Cám 6a.14%' THEN N'Cám 6a.14'
                    ELSE t.TenThan
                END AS TenThanGop,
                c.TonDK,
                c.NhapTK,
                c.XuatTK,
                c.TonCK
            FROM dbo.CDOTHAN c
            LEFT JOIN TTCO_QTHT.dbo.DMKHO k
                ON UPPER(LTRIM(RTRIM(c.MaKho))) = UPPER(LTRIM(RTRIM(k.MaKho)))
            LEFT JOIN TTCO_QTHT.dbo.DMTHAN t
                ON UPPER(LTRIM(RTRIM(c.MaThan))) = UPPER(LTRIM(RTRIM(t.MaThan)))
            WHERE c.NamHT = ?
              AND c.ThangHT = ?
              AND (
                  t.TenThan LIKE N'Cám 5a.10%'
                  OR t.TenThan LIKE N'Cám 5a.14%'
                  OR t.TenThan LIKE N'Cám 6a.10%'
                  OR t.TenThan LIKE N'Cám 6a.14%'
              )
        )
        SELECT
            NamHT,
            ThangHT,
            MaKho,
            TenKho,
            TenThanGop,
            CAST(SUM(ISNULL(TonDK, 0)) AS float) AS TonDauKy,
            CAST(SUM(ISNULL(NhapTK, 0)) AS float) AS NhapTrongKy,
            CAST(SUM(ISNULL(XuatTK, 0)) AS float) AS XuatTrongKy,
            CAST(SUM(ISNULL(TonCK, 0)) AS float) AS TonCuoiKy,
            COUNT(*) AS SoDongGop
        FROM src
        GROUP BY
            NamHT,
            ThangHT,
            MaKho,
            TenKho,
            TenThanGop
        ORDER BY
            TenKho,
            TenThanGop
        """
        rows = run_query(conn, sql_fallback, [nam, thang])
        for r in rows:
            r["GhiChu"] = f"SQL Server không hỗ trợ STRING_AGG hoặc lỗi STRING_AGG: {exc}"
        return rows


def get_all_columns_across_dbs(config: Dict[str, str]) -> List[Dict[str, Any]]:
    rows = []
    for db in DB_NAMES:
        try:
            with pyodbc.connect(build_conn_str(config, db), timeout=30) as conn:
                sql = """
                SELECT
                    DB_NAME() AS DatabaseName,
                    TABLE_SCHEMA,
                    TABLE_NAME,
                    COLUMN_NAME,
                    DATA_TYPE,
                    ORDINAL_POSITION
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA NOT IN ('sys', 'INFORMATION_SCHEMA')
                ORDER BY TABLE_SCHEMA, TABLE_NAME, ORDINAL_POSITION
                """
                rows.extend(run_query(conn, sql))
        except Exception as exc:
            rows.append({
                "DatabaseName": db,
                "TABLE_SCHEMA": "",
                "TABLE_NAME": "",
                "COLUMN_NAME": "",
                "DATA_TYPE": "",
                "ORDINAL_POSITION": "",
                "ERROR": str(exc),
            })
    return rows


def candidate_group_columns(columns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    key_terms = [
        "mathan", "tenthan", "nhom", "loai", "cha", "cap", "group", "parent",
        "masp", "tensp", "mact", "tenct", "mahang", "tenhang",
        "bc05", "cdothan"
    ]

    grouped: Dict[Tuple[str, str, str], Dict[str, Any]] = {}

    for c in columns:
        col = str(c.get("COLUMN_NAME", "")).lower()
        table = str(c.get("TABLE_NAME", "")).lower()
        hit = any(term in col or term in table for term in key_terms)
        if not hit:
            continue

        key = (c.get("DatabaseName", ""), c.get("TABLE_SCHEMA", ""), c.get("TABLE_NAME", ""))
        item = grouped.setdefault(key, {
            "DatabaseName": c.get("DatabaseName", ""),
            "TABLE_SCHEMA": c.get("TABLE_SCHEMA", ""),
            "TABLE_NAME": c.get("TABLE_NAME", ""),
            "COLUMNS": [],
            "COLUMN_COUNT": 0,
            "HAS_MATHAN": "NO",
            "HAS_TENTHAN": "NO",
            "HAS_NHOM_LOAI_CHA_CAP": "NO",
        })

        col_name = c.get("COLUMN_NAME", "")
        item["COLUMNS"].append(f"{col_name}({c.get('DATA_TYPE', '')})")
        item["COLUMN_COUNT"] += 1

        if col == "mathan":
            item["HAS_MATHAN"] = "YES"
        if "tenthan" in col or ("ten" in col and "than" in col):
            item["HAS_TENTHAN"] = "YES"
        if any(term in col for term in ["nhom", "loai", "cha", "cap", "group", "parent"]):
            item["HAS_NHOM_LOAI_CHA_CAP"] = "YES"

    result = list(grouped.values())
    for r in result:
        r["COLUMNS"] = "; ".join(r["COLUMNS"])
        score = 0
        score += 100 if r["HAS_MATHAN"] == "YES" else 0
        score += 70 if r["HAS_TENTHAN"] == "YES" else 0
        score += 50 if r["HAS_NHOM_LOAI_CHA_CAP"] == "YES" else 0
        if "dm" in str(r["TABLE_NAME"]).lower():
            score += 20
        r["SCORE"] = score

    result.sort(key=lambda x: (-x["SCORE"], x["DatabaseName"], x["TABLE_NAME"]))
    return result


def get_views_and_procs(config: Dict[str, str]) -> List[Dict[str, Any]]:
    result = []
    search_terms = ["BC05", "CDOTHAN", "MaThan", "TenThan", "DMTHAN", "MaNhom", "MaLoai"]

    for db in DB_NAMES:
        try:
            with pyodbc.connect(build_conn_str(config, db), timeout=30) as conn:
                for term in search_terms:
                    sql = """
                    SELECT
                        DB_NAME() AS DatabaseName,
                        o.type_desc,
                        s.name AS SchemaName,
                        o.name AS ObjectName,
                        ? AS SearchTerm,
                        CASE
                            WHEN m.definition IS NULL THEN ''
                            ELSE SUBSTRING(m.definition, 1, 3900)
                        END AS DefinitionStart
                    FROM sys.objects o
                    INNER JOIN sys.schemas s ON o.schema_id = s.schema_id
                    LEFT JOIN sys.sql_modules m ON o.object_id = m.object_id
                    WHERE
                        o.type IN ('P', 'V', 'FN', 'IF', 'TF')
                        AND (
                            o.name LIKE ?
                            OR m.definition LIKE ?
                        )
                    ORDER BY o.type_desc, s.name, o.name
                    """
                    rows = run_query(conn, sql, [term, f"%{term}%", f"%{term}%"])
                    result.extend(rows)
        except Exception as exc:
            result.append({
                "DatabaseName": db,
                "type_desc": "",
                "SchemaName": "",
                "ObjectName": "",
                "SearchTerm": "",
                "DefinitionStart": "",
                "ERROR": str(exc),
            })

    # Deduplicate
    seen = set()
    deduped = []
    for r in result:
        key = (
            r.get("DatabaseName"),
            r.get("type_desc"),
            r.get("SchemaName"),
            r.get("ObjectName"),
            r.get("SearchTerm"),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(r)

    return deduped


def get_sample_from_candidate_tables(config: Dict[str, str], candidate_tables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out = []
    for item in candidate_tables[:40]:
        db = item.get("DatabaseName", "")
        schema = item.get("TABLE_SCHEMA", "")
        table = item.get("TABLE_NAME", "")
        if not db or not schema or not table:
            continue

        columns_text = item.get("COLUMNS", "")
        # Chỉ lấy mẫu bảng có MaThan hoặc tên bảng/cột đáng nghi.
        if item.get("HAS_MATHAN") != "YES" and item.get("SCORE", 0) < 100:
            continue

        try:
            with pyodbc.connect(build_conn_str(config, db), timeout=30) as conn:
                sql = f"SELECT TOP 50 * FROM {qident(schema)}.{qident(table)}"
                rows = run_query(conn, sql)
                for r in rows:
                    sample = {
                        "DatabaseName": db,
                        "TABLE_SCHEMA": schema,
                        "TABLE_NAME": table,
                        "SCORE": item.get("SCORE", 0),
                    }
                    sample.update(r)
                    out.append(sample)
        except Exception as exc:
            out.append({
                "DatabaseName": db,
                "TABLE_SCHEMA": schema,
                "TABLE_NAME": table,
                "ERROR": str(exc),
                "COLUMNS": columns_text,
            })

    return out


def add_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])

    if not rows:
        ws.append(["Không có dữ liệu"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, min(ws.max_row, 300) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def readme_rows(config: Dict[str, str], nam: int, thang: int) -> List[Dict[str, Any]]:
    return [
        {"Muc": "Mục tiêu", "NoiDung": "Tìm cơ chế TTCOAPP dùng để tổng hợp Cám 5a.10, Cám 5a.14, Cám 6a.10, Cám 6a.14."},
        {"Muc": "Không dùng mapping Excel", "NoiDung": "File này chỉ kiểm tra trực tiếp database."},
        {"Muc": "Database quét", "NoiDung": ", ".join(DB_NAMES)},
        {"Muc": "Kỳ CDOTHAN", "NoiDung": f"{thang}/{nam}"},
        {"Muc": "Sheet DMTHAN_COLUMNS", "NoiDung": "Toàn bộ cột của TTCO_QTHT.dbo.DMTHAN. Kiểm tra có MaNhom/MaLoai/MaCha/Cap không."},
        {"Muc": "Sheet DMTHAN_TARGET_ROWS", "NoiDung": "Các dòng DMTHAN chứa 4 nhóm Cám cần kiểm tra."},
        {"Muc": "Sheet CDOTHAN_TARGET_DETAIL", "NoiDung": "Tồn kho chi tiết theo các mã than thuộc 4 nhóm Cám."},
        {"Muc": "Sheet CDOTHAN_TARGET_GROUPED_BASE", "NoiDung": "Tồn kho được gộp thử theo tên gốc Cám 5a.10/5a.14/6a.10/6a.14 để đối chiếu."},
        {"Muc": "Sheet CANDIDATE_GROUP_TABLES", "NoiDung": "Các bảng/cột có khả năng chứa cơ chế nhóm than."},
        {"Muc": "Sheet OBJECTS_SEARCH", "NoiDung": "Stored procedure/view/function có liên quan BC05, CDOTHAN, MaThan, TenThan, DMTHAN..."},
        {"Muc": "Sheet CANDIDATE_TABLE_SAMPLES", "NoiDung": "Mẫu dữ liệu từ bảng ứng viên để xác định bảng nhóm chuẩn."},
    ]


def main() -> int:
    print("=" * 70)
    print("TTCO CHECK CO CHE TONG HOP CHUNG LOAI THAN")
    print("=" * 70)

    config = load_config()
    print("DB_SERVER:", config["DB_SERVER"])
    print("DB_AUTH:", config["DB_AUTH"])
    print()

    with pyodbc.connect(build_conn_str(config, "TTCO_THONGKE"), timeout=30) as conn_thongke:
        nam, thang = get_latest_period(conn_thongke)
        cdothan_detail = get_cdothan_target_rows(conn_thongke, nam, thang)
        cdothan_grouped = get_cdothan_target_grouped_by_base(conn_thongke, nam, thang)

    with pyodbc.connect(build_conn_str(config, "TTCO_QTHT"), timeout=30) as conn_qtht:
        dmthan_columns = get_dmthan_columns(conn_qtht)
        dmthan_all = get_dmthan_all(conn_qtht)
        dmthan_target = get_dmthan_target_rows(conn_qtht)

    all_columns = get_all_columns_across_dbs(config)
    candidate_tables = candidate_group_columns(all_columns)
    objects = get_views_and_procs(config)
    samples = get_sample_from_candidate_tables(config, candidate_tables)

    wb = Workbook()
    ws0 = wb.active
    wb.remove(ws0)

    add_sheet(wb, "README", readme_rows(config, nam, thang))
    add_sheet(wb, "DMTHAN_COLUMNS", dmthan_columns)
    add_sheet(wb, "DMTHAN_TARGET_ROWS", dmthan_target)
    add_sheet(wb, "CDOTHAN_TARGET_DETAIL", cdothan_detail)
    add_sheet(wb, "CDOTHAN_TARGET_GROUPED_BASE", cdothan_grouped)
    add_sheet(wb, "CANDIDATE_GROUP_TABLES", candidate_tables)
    add_sheet(wb, "CANDIDATE_TABLE_SAMPLES", samples)
    add_sheet(wb, "OBJECTS_SEARCH", objects)
    add_sheet(wb, "DMTHAN_ALL", dmthan_all)
    add_sheet(wb, "ALL_COLUMNS", all_columns)

    output = OUTPUT_DIR / f"TTCO_CHECK_CO_CHE_TONG_HOP_THAN_{nam}_{thang:02d}.xlsx"
    wb.save(output)

    print("Đã tạo file:")
    print(output.resolve())
    print()
    print("Gửi lại file này để kiểm tra cơ chế nhóm than chuẩn.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print()
        print("LỖI:", exc)
        print()
        raise SystemExit(1)
