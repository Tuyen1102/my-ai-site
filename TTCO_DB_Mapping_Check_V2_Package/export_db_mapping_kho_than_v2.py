import json
import os
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pyodbc
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("mapping_output")
OUTPUT_DIR.mkdir(exist_ok=True)

ACTIVE_DB = "TTCO_THONGKE"
DEFAULT_DATABASES = ["TTCO_THONGKE", "TTCO_QTHT", "TTCO_DHSX"]

THAN_CODE_COLUMN_KEYWORDS = ["mathan", "mact", "masp", "mahang", "maloai", "maloaithan", "ma"]
KHO_CODE_COLUMN_KEYWORDS = ["makho", "makho_n", "makho_x", "makho_nhan", "makho_nghg", "kho", "madiem", "ma"]
NAME_COLUMN_KEYWORDS = ["ten", "tenct", "tenthan", "tenkho", "tieude", "diengiai", "mota", "ghichu", "name"]


def norm(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")

    return value


def qident(name: str) -> str:
    return "[" + str(name).replace("]", "]]") + "]"


def load_config() -> Dict[str, Any]:
    load_dotenv()

    db_list_raw = os.getenv("SCAN_DATABASES", ",".join(DEFAULT_DATABASES)).strip()
    db_list = [x.strip() for x in db_list_raw.split(",") if x.strip()]
    if not db_list:
        db_list = DEFAULT_DATABASES

    return {
        "DB_SERVER": os.getenv("DB_SERVER", "tuyenthancuaong.com.vn,52376").strip(),
        "DB_DRIVER": os.getenv("DB_DRIVER", "ODBC Driver 18 for SQL Server").strip(),
        "DB_AUTH": os.getenv("DB_AUTH", "windows").strip().lower(),
        "DB_USER": os.getenv("DB_USER", "").strip(),
        "DB_PASSWORD": os.getenv("DB_PASSWORD", "").strip(),
        "SCAN_DATABASES": db_list,
    }


def build_conn_str(config: Dict[str, Any], database: str) -> str:
    base = (
        f"DRIVER={{{config['DB_DRIVER']}}};"
        f"SERVER={config['DB_SERVER']};"
        f"DATABASE={database};"
        "TrustServerCertificate=yes;"
        "Encrypt=no;"
    )

    if config["DB_AUTH"] == "sql":
        if not config["DB_USER"] or not config["DB_PASSWORD"]:
            raise RuntimeError("DB_AUTH=sql nhưng thiếu DB_USER hoặc DB_PASSWORD trong file .env")
        return base + f"UID={config['DB_USER']};PWD={config['DB_PASSWORD']};"

    return base + "Trusted_Connection=yes;"


def connect(config: Dict[str, Any], database: str) -> pyodbc.Connection:
    return pyodbc.connect(build_conn_str(config, database), timeout=30)


def rows_to_dicts(cursor: pyodbc.Cursor) -> List[Dict[str, Any]]:
    cols = [c[0] for c in cursor.description]
    out: List[Dict[str, Any]] = []
    for row in cursor.fetchall():
        out.append({cols[i]: norm(row[i]) for i in range(len(cols))})
    return out


def run_query(conn: pyodbc.Connection, sql: str, params: Optional[List[Any]] = None) -> List[Dict[str, Any]]:
    cur = conn.cursor()
    cur.execute(sql, params or [])
    return rows_to_dicts(cur)


def is_text_type(data_type: str) -> bool:
    return str(data_type).lower() in {"varchar", "nvarchar", "char", "nchar", "text", "ntext"}


def get_db_columns(config: Dict[str, Any], database: str) -> List[Dict[str, Any]]:
    with connect(config, database) as conn:
        sql = """
        SELECT
            TABLE_CATALOG,
            TABLE_SCHEMA,
            TABLE_NAME,
            COLUMN_NAME,
            DATA_TYPE,
            ORDINAL_POSITION
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA NOT IN ('sys', 'INFORMATION_SCHEMA')
        ORDER BY TABLE_SCHEMA, TABLE_NAME, ORDINAL_POSITION
        """
        return run_query(conn, sql)


def get_latest_period_and_stock(config: Dict[str, Any]) -> Tuple[int, int, List[Dict[str, Any]]]:
    with connect(config, ACTIVE_DB) as conn:
        latest_sql = """
        SELECT TOP 1 NamHT, ThangHT
        FROM CDOTHAN
        WHERE NamHT IS NOT NULL AND ThangHT IS NOT NULL
        GROUP BY NamHT, ThangHT
        ORDER BY NamHT DESC, ThangHT DESC
        """
        latest = run_query(conn, latest_sql)
        if not latest:
            raise RuntimeError("Không tìm thấy kỳ dữ liệu trong TTCO_THONGKE.dbo.CDOTHAN")

        nam = int(latest[0]["NamHT"])
        thang = int(latest[0]["ThangHT"])

        stock_sql = """
        SELECT
            NamHT,
            ThangHT,
            LTRIM(RTRIM(MaKho)) AS MaKho,
            LTRIM(RTRIM(MaThan)) AS MaThan,
            CAST(SUM(ISNULL(TonDK, 0)) AS float) AS TonDauKy,
            CAST(SUM(ISNULL(NhapTK, 0)) AS float) AS NhapTrongKy,
            CAST(SUM(ISNULL(XuatTK, 0)) AS float) AS XuatTrongKy,
            CAST(SUM(ISNULL(TonCK, 0)) AS float) AS TonCuoiKy
        FROM CDOTHAN
        WHERE NamHT = ?
          AND ThangHT = ?
          AND MaKho IS NOT NULL
          AND MaThan IS NOT NULL
          AND LTRIM(RTRIM(MaKho)) <> ''
          AND LTRIM(RTRIM(MaThan)) <> ''
        GROUP BY
            NamHT,
            ThangHT,
            LTRIM(RTRIM(MaKho)),
            LTRIM(RTRIM(MaThan))
        HAVING SUM(ISNULL(TonCK, 0)) <> 0
        ORDER BY LTRIM(RTRIM(MaKho)), LTRIM(RTRIM(MaThan))
        """
        stock = run_query(conn, stock_sql, [nam, thang])
        return nam, thang, stock


def chunked(values: Sequence[str], size: int = 200) -> Iterable[List[str]]:
    values = list(values)
    for i in range(0, len(values), size):
        yield values[i:i + size]


def column_may_hold_codes(column_name: str, target: str) -> bool:
    n = column_name.lower()

    if target == "than":
        return any(k in n for k in THAN_CODE_COLUMN_KEYWORDS)

    if target == "kho":
        return any(k in n for k in KHO_CODE_COLUMN_KEYWORDS)

    return False


def score_name_column(column_name: str, target: str) -> int:
    n = column_name.lower()
    score = 0

    if "ten" in n:
        score += 50
    if "name" in n:
        score += 30
    if "tieude" in n or "tieu_de" in n:
        score += 25
    if "diengiai" in n or "dien_giai" in n or "mota" in n or "mo_ta" in n:
        score += 20
    if "ghichu" in n or "ghi_chu" in n:
        score += 8

    if target == "than":
        if "than" in n:
            score += 50
        if "ct" in n:
            score += 15
    elif target == "kho":
        if "kho" in n:
            score += 50
        if "diem" in n:
            score += 15

    if n.startswith("ma") or n in {"mathan", "makho", "mact", "masp"}:
        score -= 100
    if "login" in n or "user" in n or "pass" in n:
        score -= 100

    return score


def get_table_columns_by_table(all_columns: List[Dict[str, Any]]) -> Dict[Tuple[str, str, str], List[Dict[str, Any]]]:
    out: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = {}
    for col in all_columns:
        key = (col["TABLE_CATALOG"], col["TABLE_SCHEMA"], col["TABLE_NAME"])
        out.setdefault(key, []).append(col)
    return out


def make_in_clause_params(values: Sequence[str]) -> Tuple[str, List[str]]:
    placeholders = ",".join(["?"] * len(values))
    return placeholders, list(values)


def scan_exact_code_matches(
    config: Dict[str, Any],
    database: str,
    columns_by_table: Dict[Tuple[str, str, str], List[Dict[str, Any]]],
    active_codes: Sequence[str],
    target: str,
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    active_codes = sorted(set(str(x).strip() for x in active_codes if str(x).strip()))
    if not active_codes:
        return out

    with connect(config, database) as conn:
        for (catalog, schema, table), cols in columns_by_table.items():
            if catalog != database:
                continue

            text_cols = [c for c in cols if is_text_type(c["DATA_TYPE"])]
            code_cols = [c for c in text_cols if column_may_hold_codes(c["COLUMN_NAME"], target)]
            if not code_cols:
                continue

            candidate_name_cols = sorted(
                [c for c in text_cols if score_name_column(c["COLUMN_NAME"], target) > 0],
                key=lambda c: score_name_column(c["COLUMN_NAME"], target),
                reverse=True,
            )

            for code_col in code_cols:
                code_col_name = code_col["COLUMN_NAME"]

                for chunk in chunked(active_codes, 150):
                    placeholders, params = make_in_clause_params(chunk)

                    # Lấy sample toàn bộ text columns ưu tiên để người dùng xem dòng nào chứa mã.
                    sample_cols = [code_col_name]
                    for c in candidate_name_cols[:8]:
                        if c["COLUMN_NAME"] not in sample_cols:
                            sample_cols.append(c["COLUMN_NAME"])

                    # Nếu chưa có cột tên, vẫn lấy thêm vài text col để soi.
                    if len(sample_cols) < 6:
                        for c in text_cols:
                            if c["COLUMN_NAME"] not in sample_cols:
                                sample_cols.append(c["COLUMN_NAME"])
                            if len(sample_cols) >= 8:
                                break

                    select_parts = [
                        f"LTRIM(RTRIM(CAST({qident(c)} AS nvarchar(4000)))) AS {qident(c)}"
                        for c in sample_cols
                    ]
                    select_sql = ", ".join(select_parts)

                    sql = f"""
                    SELECT DISTINCT TOP 500
                        {select_sql}
                    FROM {qident(schema)}.{qident(table)}
                    WHERE LTRIM(RTRIM(CAST({qident(code_col_name)} AS nvarchar(4000)))) IN ({placeholders})
                    ORDER BY {qident(code_col_name)}
                    """

                    try:
                        rows = run_query(conn, sql, params)
                    except Exception as exc:
                        out.append({
                            "Database": database,
                            "Target": target,
                            "Table": f"{schema}.{table}",
                            "CodeColumn": code_col_name,
                            "Error": str(exc),
                        })
                        continue

                    for row in rows:
                        code_value = str(row.get(code_col_name, "")).strip()
                        best_col = ""
                        best_val = ""
                        best_score = -999
                        for c in candidate_name_cols:
                            col_name = c["COLUMN_NAME"]
                            value = str(row.get(col_name, "") or "").strip()
                            sc = score_name_column(col_name, target)
                            if value and value.lower() != code_value.lower() and sc > best_score:
                                best_col = col_name
                                best_val = value
                                best_score = sc

                        out.append({
                            "Database": database,
                            "Target": target,
                            "Table": f"{schema}.{table}",
                            "CodeColumn": code_col_name,
                            "CodeValue": code_value,
                            "BestNameColumn": best_col,
                            "BestNameValue": best_val,
                            "NameScore": best_score if best_col else "",
                            "SampleRowJSON": json.dumps(row, ensure_ascii=False),
                        })

    return out


def build_best_mapping(rows: List[Dict[str, Any]], target: str) -> Dict[str, str]:
    # Chọn tên tốt nhất theo điểm cột tên, ưu tiên bảng có vẻ danh mục/tmp, loại trừ transaction/log/user.
    best: Dict[str, Tuple[int, str, Dict[str, Any]]] = {}

    for r in rows:
        code = str(r.get("CodeValue", "") or "").strip()
        name = str(r.get("BestNameValue", "") or "").strip()
        table = str(r.get("Table", "") or "").lower()
        try:
            score = int(r.get("NameScore") or 0)
        except Exception:
            score = 0

        if not code or not name:
            continue
        if name.lower() == code.lower():
            continue
        if name.lower() in {"auto", "doanh"}:
            continue
        if "tenlogin" in str(r.get("BestNameColumn", "")).lower():
            continue

        bonus = 0
        if "dm" in table or "danh" in table:
            bonus += 80
        if "tmp" in table:
            bonus += 20
        if "than" in table and target == "than":
            bonus += 10
        if "kho" in table and target == "kho":
            bonus += 10

        penalty = 0
        if table.startswith("dbo.than") or ".than" in table:
            penalty += 20
        if "log" in table or "nsd" in table or "user" in table:
            penalty += 100

        final_score = score + bonus - penalty

        if code not in best or final_score > best[code][0]:
            best[code] = (final_score, name, r)

    return {code: value[1] for code, value in best.items()}


def make_final_review_rows(
    stock_rows: List[Dict[str, Any]],
    kho_mapping: Dict[str, str],
    than_mapping: Dict[str, str],
) -> List[Dict[str, Any]]:
    out = []
    for r in stock_rows:
        ma_kho = str(r.get("MaKho", "") or "").strip()
        ma_than = str(r.get("MaThan", "") or "").strip()
        ten_kho = kho_mapping.get(ma_kho, "")
        ten_than = than_mapping.get(ma_than, "")
        item = dict(r)
        item["TenKho_DB"] = ten_kho
        item["TenThan_DB"] = ten_than
        item["KiemTraKho"] = "OK" if ten_kho else "THIEU_TEN_KHO"
        item["KiemTraThan"] = "OK" if ten_than else "THIEU_TEN_THAN"
        out.append(item)
    return out


def add_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])

    if not rows:
        ws.append(["Không có dữ liệu"])
        ws.freeze_panes = "A2"
        return

    headers = list(rows[0].keys())
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, min(ws.max_row, 300) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> int:
    config = load_config()

    print("=" * 80)
    print("TTCO DB MAPPING CHECK V2 - SCAN ALL DATABASES")
    print("=" * 80)
    print("DB_SERVER:", config["DB_SERVER"])
    print("DB_AUTH:", config["DB_AUTH"])
    print("SCAN_DATABASES:", ", ".join(config["SCAN_DATABASES"]))
    print()

    nam, thang, stock_rows = get_latest_period_and_stock(config)
    active_kho_codes = sorted({str(r["MaKho"]).strip() for r in stock_rows if str(r.get("MaKho", "")).strip()})
    active_than_codes = sorted({str(r["MaThan"]).strip() for r in stock_rows if str(r.get("MaThan", "")).strip()})

    print(f"Kỳ tồn kho: {thang}/{nam}")
    print(f"Số dòng tồn kho: {len(stock_rows)}")
    print(f"Số mã kho cần dò: {len(active_kho_codes)}")
    print(f"Số mã than cần dò: {len(active_than_codes)}")
    print()

    all_columns: List[Dict[str, Any]] = []
    than_matches: List[Dict[str, Any]] = []
    kho_matches: List[Dict[str, Any]] = []

    for database in config["SCAN_DATABASES"]:
        print(f"Đang đọc cấu trúc DB: {database}")
        try:
            db_cols = get_db_columns(config, database)
        except Exception as exc:
            print(f"  LỖI đọc cấu trúc {database}: {exc}")
            all_columns.append({"TABLE_CATALOG": database, "ERROR": str(exc)})
            continue

        all_columns.extend(db_cols)
        cols_by_table = get_table_columns_by_table(db_cols)

        print(f"  Đang dò mã chủng loại than trong {database}...")
        than_matches.extend(scan_exact_code_matches(config, database, cols_by_table, active_than_codes, "than"))

        print(f"  Đang dò mã kho trong {database}...")
        kho_matches.extend(scan_exact_code_matches(config, database, cols_by_table, active_kho_codes, "kho"))

    than_map = build_best_mapping(than_matches, "than")
    kho_map = build_best_mapping(kho_matches, "kho")
    final_review = make_final_review_rows(stock_rows, kho_map, than_map)

    readme = [
        {"Muc": "Mục tiêu", "NoiDung": "Dò tên kho/tên chủng loại than trực tiếp trong các database, không dùng file Excel báo cáo."},
        {"Muc": "Database đã quét", "NoiDung": ", ".join(config["SCAN_DATABASES"])},
        {"Muc": "Nguồn tồn kho", "NoiDung": "TTCO_THONGKE.dbo.CDOTHAN"},
        {"Muc": "Kỳ dữ liệu", "NoiDung": f"{thang}/{nam}"},
        {"Muc": "Sheet cần kiểm tra trước", "NoiDung": "TON_KHO_FINAL_REVIEW"},
        {"Muc": "Nếu TenThan_DB còn trống", "NoiDung": "Xem SCAN_THAN_CODE_MATCHES để xác định bảng/cột nào chứa tên đúng."},
        {"Muc": "Nếu TenKho_DB còn trống", "NoiDung": "Xem SCAN_KHO_CODE_MATCHES để xác định bảng/cột nào chứa tên đúng."},
    ]

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "README", readme)
    add_sheet(wb, "TON_KHO_FINAL_REVIEW", final_review)
    add_sheet(wb, "SCAN_THAN_CODE_MATCHES", than_matches)
    add_sheet(wb, "SCAN_KHO_CODE_MATCHES", kho_matches)
    add_sheet(wb, "TON_KHO_CODE_GOC", stock_rows)
    add_sheet(wb, "ALL_TABLE_COLUMNS", all_columns)

    output = OUTPUT_DIR / f"TTCO_DB_MAPPING_KHO_THAN_V2_{nam}_{thang:02d}.xlsx"
    wb.save(output)

    print()
    print("Đã tạo file:")
    print(output.resolve())
    print()
    print("Anh gửi lại file này để tôi chốt bảng/cột danh mục chính xác.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print()
        print("LỖI:", exc)
        print()
        raise SystemExit(1)
