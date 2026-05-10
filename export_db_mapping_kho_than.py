import os
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pyodbc
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("mapping_output")
OUTPUT_DIR.mkdir(exist_ok=True)


def norm(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return value


def qident(name: str) -> str:
    return "[" + name.replace("]", "]]") + "]"


def load_config() -> Dict[str, str]:
    load_dotenv()
    return {
        "DB_SERVER": os.getenv("DB_SERVER", "tuyenthancuaong.com.vn,52376").strip(),
        "DB_NAME": os.getenv("DB_NAME", "TTCO_THONGKE").strip(),
        "DB_DRIVER": os.getenv("DB_DRIVER", "ODBC Driver 18 for SQL Server").strip(),
        "DB_AUTH": os.getenv("DB_AUTH", "windows").strip().lower(),
        "DB_USER": os.getenv("DB_USER", "").strip(),
        "DB_PASSWORD": os.getenv("DB_PASSWORD", "").strip(),
    }


def build_conn_str(config: Dict[str, str]) -> str:
    base = (
        f"DRIVER={{{config['DB_DRIVER']}}};"
        f"SERVER={config['DB_SERVER']};"
        f"DATABASE={config['DB_NAME']};"
        "TrustServerCertificate=yes;"
        "Encrypt=no;"
    )

    if config["DB_AUTH"] == "sql":
        if not config["DB_USER"] or not config["DB_PASSWORD"]:
            raise RuntimeError("DB_AUTH=sql nhưng thiếu DB_USER hoặc DB_PASSWORD trong file .env")
        return base + f"UID={config['DB_USER']};PWD={config['DB_PASSWORD']};"

    return base + "Trusted_Connection=yes;"


def rows_to_dicts(cursor: pyodbc.Cursor) -> List[Dict[str, Any]]:
    cols = [c[0] for c in cursor.description]
    out: List[Dict[str, Any]] = []
    for row in cursor.fetchall():
        out.append({cols[i]: norm(row[i]) for i in range(len(cols))})
    return out


def run_query(conn: pyodbc.Connection, sql: str, params: List[Any] | None = None) -> List[Dict[str, Any]]:
    cur = conn.cursor()
    cur.execute(sql, params or [])
    return rows_to_dicts(cur)


def get_columns(conn: pyodbc.Connection) -> List[Dict[str, Any]]:
    sql = """
    SELECT
        TABLE_SCHEMA,
        TABLE_NAME,
        COLUMN_NAME,
        DATA_TYPE,
        ORDINAL_POSITION
    FROM INFORMATION_SCHEMA.COLUMNS
    WHERE TABLE_SCHEMA NOT IN ('sys', 'INFORMATION_SCHEMA')
    ORDER BY TABLE_SCHEMA, TABLE_NAME, ORDINAL_POSITION
    """
    return run_query(conn, sql)


def get_latest_period(conn: pyodbc.Connection) -> Tuple[int, int]:
    sql = """
    SELECT TOP 1 NamHT, ThangHT
    FROM CDOTHAN
    WHERE NamHT IS NOT NULL AND ThangHT IS NOT NULL
    GROUP BY NamHT, ThangHT
    ORDER BY NamHT DESC, ThangHT DESC
    """
    data = run_query(conn, sql)
    if not data:
        raise RuntimeError("Không tìm thấy kỳ dữ liệu trong bảng CDOTHAN")
    return int(data[0]["NamHT"]), int(data[0]["ThangHT"])


def get_ton_kho_codes(conn: pyodbc.Connection, nam: int, thang: int) -> List[Dict[str, Any]]:
    sql = """
    SELECT
        NamHT,
        ThangHT,
        LTRIM(RTRIM(MaKho)) AS MaKho,
        LTRIM(RTRIM(MaThan)) AS MaThan,
        CAST(SUM(ISNULL(TonDK, 0)) AS float) AS TonDauKy,
        CAST(SUM(ISNULL(NhapTK, 0)) AS float) AS NhapTrongKy,
        CAST(SUM(ISNULL(XuatTK, 0)) AS float) AS XuatTrongKy,
        CAST(SUM(ISNULL(TonCK, 0)) AS float) AS TonCuoiKy
    FROM CDOTHAN
    WHERE NamHT = ?
      AND ThangHT = ?
      AND MaKho IS NOT NULL
      AND MaThan IS NOT NULL
      AND LTRIM(RTRIM(MaKho)) <> ''
      AND LTRIM(RTRIM(MaThan)) <> ''
    GROUP BY
        NamHT,
        ThangHT,
        LTRIM(RTRIM(MaKho)),
        LTRIM(RTRIM(MaThan))
    HAVING SUM(ISNULL(TonCK, 0)) <> 0
    ORDER BY LTRIM(RTRIM(MaKho)), LTRIM(RTRIM(MaThan))
    """
    return run_query(conn, sql, [nam, thang])


def group_columns(cols: List[Dict[str, Any]]) -> Dict[Tuple[str, str], List[Dict[str, Any]]]:
    grouped: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for c in cols:
        grouped.setdefault((c["TABLE_SCHEMA"], c["TABLE_NAME"]), []).append(c)
    return grouped


def is_text_type(data_type: str) -> bool:
    return data_type.lower() in {"varchar", "nvarchar", "char", "nchar", "text", "ntext"}


def score_name_column(column_name: str, target: str) -> int:
    name = column_name.lower()
    score = 0

    if target == "than":
        if name in {"tenthan", "ten_than", "tenloai", "tenloaithan", "loai_than", "loaithan"}:
            score += 100
        if "than" in name:
            score += 40
        if "ten" in name:
            score += 30
        if "name" in name:
            score += 20
        if "mota" in name or "diengiai" in name or "ghichu" in name:
            score += 10

    if target == "kho":
        if name in {"tenkho", "ten_kho", "tenkhohang", "khohang"}:
            score += 100
        if "kho" in name:
            score += 40
        if "ten" in name:
            score += 30
        if "name" in name:
            score += 20
        if "mota" in name or "diengiai" in name or "ghichu" in name:
            score += 10

    if name.startswith("ma"):
        score -= 50
    if name in {"mathan", "makho"}:
        score -= 100

    return score


def find_candidate_tables(cols_by_table: Dict[Tuple[str, str], List[Dict[str, Any]]], code_col: str, target: str) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []

    for (schema, table), cols in cols_by_table.items():
        col_names = {c["COLUMN_NAME"].lower(): c["COLUMN_NAME"] for c in cols}
        if code_col.lower() not in col_names:
            continue

        text_cols = [
            c["COLUMN_NAME"] for c in cols
            if is_text_type(c["DATA_TYPE"]) and c["COLUMN_NAME"].lower() != code_col.lower()
        ]

        scored = [(score_name_column(c, target), c) for c in text_cols]
        scored = [(s, c) for s, c in scored if s > 0]
        scored.sort(reverse=True)

        candidates.append({
            "TABLE_SCHEMA": schema,
            "TABLE_NAME": table,
            "CODE_COLUMN": col_names[code_col.lower()],
            "BEST_NAME_COLUMN": scored[0][1] if scored else "",
            "NAME_COLUMN_SCORE": scored[0][0] if scored else 0,
            "TEXT_COLUMNS": ", ".join(text_cols),
            "SCORED_NAME_COLUMNS": ", ".join([f"{c}({s})" for s, c in scored[:8]]),
        })

    candidates.sort(key=lambda x: (-x["NAME_COLUMN_SCORE"], x["TABLE_NAME"]))
    return candidates


def fetch_candidate_mapping_rows(
    conn: pyodbc.Connection,
    candidates: List[Dict[str, Any]],
    code_col_expected: str,
    target: str,
    active_codes: set[str],
    max_tables: int = 20,
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []

    for cand in candidates[:max_tables]:
        schema = cand["TABLE_SCHEMA"]
        table = cand["TABLE_NAME"]
        code_col = cand["CODE_COLUMN"]
        text_cols = [c.strip() for c in cand["TEXT_COLUMNS"].split(",") if c.strip()]
        best_name_col = cand["BEST_NAME_COLUMN"]

        select_cols = [code_col]
        if best_name_col and best_name_col not in select_cols:
            select_cols.append(best_name_col)

        for col in text_cols:
            if col not in select_cols and len(select_cols) < 8:
                select_cols.append(col)

        select_sql = ", ".join([f"LTRIM(RTRIM(CAST({qident(c)} AS nvarchar(4000)))) AS {qident(c)}" for c in select_cols])

        sql = f"""
        SELECT DISTINCT TOP 2000
            {select_sql}
        FROM {qident(schema)}.{qident(table)}
        WHERE {qident(code_col)} IS NOT NULL
          AND LTRIM(RTRIM(CAST({qident(code_col)} AS nvarchar(4000)))) <> ''
        ORDER BY {qident(code_col)}
        """

        try:
            rows = run_query(conn, sql)
        except Exception as exc:
            out.append({
                "SOURCE_TABLE": f"{schema}.{table}",
                "ERROR": str(exc),
            })
            continue

        for row in rows:
            code = str(row.get(code_col, "")).strip()
            if active_codes and code not in active_codes:
                continue

            item = {
                "SOURCE_TABLE": f"{schema}.{table}",
                "TARGET": target,
                "CODE_COLUMN": code_col,
                "CODE_VALUE": code,
                "BEST_NAME_COLUMN": best_name_col,
                "BEST_NAME_VALUE": row.get(best_name_col, "") if best_name_col else "",
                "MATCH_ACTIVE_CODE": "YES" if code in active_codes else "NO",
            }

            for col in select_cols:
                item[col] = row.get(col, "")

            out.append(item)

    return out


def choose_best_mapping(candidate_rows: List[Dict[str, Any]]) -> Dict[str, str]:
    # Ưu tiên các dòng từ bảng có BEST_NAME_VALUE không rỗng.
    mapping: Dict[str, str] = {}

    for row in candidate_rows:
        code = str(row.get("CODE_VALUE", "")).strip()
        name = str(row.get("BEST_NAME_VALUE", "")).strip()

        if not code or not name:
            continue

        # Bỏ qua trường hợp tên trùng mã.
        if name.lower() == code.lower():
            continue

        if code not in mapping:
            mapping[code] = name

    return mapping


def make_ton_kho_with_names(
    ton_rows: List[Dict[str, Any]],
    kho_map: Dict[str, str],
    than_map: Dict[str, str],
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in ton_rows:
        item = dict(row)
        ma_kho = str(row.get("MaKho", "")).strip()
        ma_than = str(row.get("MaThan", "")).strip()
        item["TenKho_DB_Guess"] = kho_map.get(ma_kho, "")
        item["TenThan_DB_Guess"] = than_map.get(ma_than, "")
        item["CanKiemTra_TenKho"] = "OK" if item["TenKho_DB_Guess"] else "THIEU_TEN_KHO"
        item["CanKiemTra_TenThan"] = "OK" if item["TenThan_DB_Guess"] else "THIEU_TEN_THAN"
        out.append(item)
    return out


def add_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])

    if not rows:
        ws.append(["Không có dữ liệu"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, min(ws.max_row, 300) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_readme_rows(config: Dict[str, str], nam: int, thang: int, ton_count: int) -> List[Dict[str, Any]]:
    return [
        {"Muc": "Mục tiêu", "NoiDung": "Kiểm tra mapping tên kho/tên chủng loại than trực tiếp từ database, không dùng file Excel báo cáo."},
        {"Muc": "Database", "NoiDung": config["DB_NAME"]},
        {"Muc": "Server", "NoiDung": config["DB_SERVER"]},
        {"Muc": "Nguồn tồn kho", "NoiDung": "CDOTHAN"},
        {"Muc": "Kỳ dữ liệu", "NoiDung": f"{thang}/{nam}"},
        {"Muc": "Số dòng tồn kho", "NoiDung": ton_count},
        {"Muc": "Sheet TON_KHO_KEM_TEN_DB", "NoiDung": "Danh sách tồn kho đang dùng trong app, kèm tên kho/tên than DB đoán được. Anh kiểm tra sheet này trước."},
        {"Muc": "Sheet DM_THAN_CANDIDATES", "NoiDung": "Các bảng/cột trong DB có thể là danh mục chủng loại than."},
        {"Muc": "Sheet DM_KHO_CANDIDATES", "NoiDung": "Các bảng/cột trong DB có thể là danh mục kho."},
        {"Muc": "Sheet TABLE_COLUMNS", "NoiDung": "Danh sách cột toàn DB để tìm bảng danh mục chính xác nếu auto guess chưa đúng."},
    ]


def main() -> int:
    print("=" * 70)
    print("TTCO DB MAPPING CHECK - KHO / CHUNG LOAI THAN")
    print("=" * 70)

    config = load_config()
    conn_str = build_conn_str(config)

    print("DB_SERVER:", config["DB_SERVER"])
    print("DB_NAME:", config["DB_NAME"])
    print("DB_AUTH:", config["DB_AUTH"])
    print()

    with pyodbc.connect(conn_str, timeout=30) as conn:
        all_cols = get_columns(conn)
        cols_by_table = group_columns(all_cols)

        nam, thang = get_latest_period(conn)
        ton_rows = get_ton_kho_codes(conn, nam, thang)

        active_kho_codes = {str(r["MaKho"]).strip() for r in ton_rows if r.get("MaKho")}
        active_than_codes = {str(r["MaThan"]).strip() for r in ton_rows if r.get("MaThan")}

        kho_tables = find_candidate_tables(cols_by_table, "MaKho", "kho")
        than_tables = find_candidate_tables(cols_by_table, "MaThan", "than")

        kho_candidate_rows = fetch_candidate_mapping_rows(
            conn=conn,
            candidates=kho_tables,
            code_col_expected="MaKho",
            target="kho",
            active_codes=active_kho_codes,
            max_tables=30,
        )

        than_candidate_rows = fetch_candidate_mapping_rows(
            conn=conn,
            candidates=than_tables,
            code_col_expected="MaThan",
            target="than",
            active_codes=active_than_codes,
            max_tables=30,
        )

        kho_map = choose_best_mapping(kho_candidate_rows)
        than_map = choose_best_mapping(than_candidate_rows)
        ton_with_names = make_ton_kho_with_names(ton_rows, kho_map, than_map)

    wb = Workbook()
    ws0 = wb.active
    wb.remove(ws0)

    add_sheet(wb, "README", build_readme_rows(config, nam, thang, len(ton_rows)))
    add_sheet(wb, "TON_KHO_KEM_TEN_DB", ton_with_names)
    add_sheet(wb, "DM_KHO_TABLES", kho_tables)
    add_sheet(wb, "DM_THAN_TABLES", than_tables)
    add_sheet(wb, "DM_KHO_CANDIDATES", kho_candidate_rows)
    add_sheet(wb, "DM_THAN_CANDIDATES", than_candidate_rows)
    add_sheet(wb, "TON_KHO_CODE_GOC", ton_rows)
    add_sheet(wb, "TABLE_COLUMNS", all_cols)

    output_xlsx = OUTPUT_DIR / f"TTCO_DB_MAPPING_KHO_THAN_{nam}_{thang:02d}.xlsx"
    wb.save(output_xlsx)

    print("Đã tạo file kiểm tra mapping:")
    print(output_xlsx.resolve())
    print()
    print("Anh mở sheet TON_KHO_KEM_TEN_DB để kiểm tra TenKho_DB_Guess và TenThan_DB_Guess.")
    print("Nếu tên chưa đúng, gửi lại file Excel này để tôi xác định bảng danh mục DB chính xác.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print()
        print("LỖI:", exc)
        print()
        raise SystemExit(1)
